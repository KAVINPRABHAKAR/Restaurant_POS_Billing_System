import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.http import FileResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from decimal import Decimal
from django.utils import timezone

# PDF Generation imports
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors

# Local imports
from .models import MenuItem, Order, OrderItem, Category
from .utils import get_sales_analytics 

# --- ROLE CHECK HELPER ---
def is_manager(user):
    return user.is_superuser or user.groups.filter(name='Manager').exists()

# --- AUTHENTICATION ---
def login_view(request):
    if request.user.is_authenticated:
        return redirect('pos_screen')
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('pos_screen')
        else:
            messages.error(request, "Invalid Username or Password.")
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

# --- POS SYSTEM (Inventory Auto-Update + GST + Service Charge) ---
@login_required
def pos_screen(request):
    items = MenuItem.objects.all()
    categories = Category.objects.all()
    last_order_id = request.session.pop('last_order_id', None)
    
    if request.method == "POST":
        selected_item_ids = request.POST.getlist('items[]')
        quantities = request.POST.getlist('qty[]')
        discount_perc = request.POST.get('discount_percentage', '0')

        try:
            discount_perc = Decimal(discount_perc)
        except (ValueError, TypeError):
            discount_perc = Decimal('0.00')
        
        if not selected_item_ids:
            messages.error(request, "Your cart is empty!")
            return redirect('pos_screen')

        try:
            with transaction.atomic():
                subtotal = Decimal('0.00')
                # Initial temporary values, updated after loop
                order = Order.objects.create(
                    cashier=request.user,
                    subtotal=0, gst_amount=0, service_charge=0, total_amount=0
                )

                for item_id, qty in zip(selected_item_ids, quantities):
                    menu_item = MenuItem.objects.select_for_update().get(id=item_id)
                    qty = int(qty)

                    if menu_item.stock < qty:
                        raise ValueError(f"Not enough stock for {menu_item.name}.")

                    # 1. DEDUCT PURCHASED QTY
                    menu_item.stock -= qty
                    
                    # 2. AUTO-REFILL LOGIC
                    if menu_item.stock <= 0:
                        menu_item.stock = 50 
                        messages.info(request, f"Inventory Alert: {menu_item.name} auto-replenished to 50.")

                    menu_item.save()

                    line_price = menu_item.price * qty
                    subtotal += line_price

                    OrderItem.objects.create(
                        order=order, item=menu_item,
                        quantity=qty, price_at_order=menu_item.price
                    )

                # CALCULATIONS (GST 5%, Service 10%)
                gst = subtotal * Decimal('0.05')
                service_charge = subtotal * Decimal('0.10')
                total_before_discount = subtotal + gst + service_charge
                discount_amount = total_before_discount * (discount_perc / Decimal('100'))
                
                order.subtotal = subtotal
                order.gst_amount = gst
                order.service_charge = service_charge
                order.discount_amount = discount_amount
                order.total_amount = total_before_discount - discount_amount
                order.save()

                request.session['last_order_id'] = order.id
                messages.success(request, f"Order #{order.id} confirmed!")
                return redirect('pos_screen')

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect('pos_screen')

    return render(request, 'billing/pos_screen.html', {
        'items': items, 'categories': categories, 'last_order_id': last_order_id
    })

@login_required
@user_passes_test(is_manager, login_url='pos_screen')
def dashboard(request):
    analytics_data = get_sales_analytics()
    today = timezone.now().date()
    orders = Order.objects.filter(order_date__date=today).order_by('-order_date')
    return render(request, 'billing/dashboard.html', {'data': analytics_data, 'orders': orders})

@login_required
@user_passes_test(is_manager, login_url='pos_screen')
def inventory_view(request):
    items = MenuItem.objects.all().order_by('category', 'name')
    return render(request, 'billing/inventory.html', {'items': items})

@login_required
@user_passes_test(is_manager, login_url='pos_screen')
def sales_report(request):
    orders = Order.objects.all().order_by('-order_date')
    return render(request, 'billing/sales_report.html', {'orders': orders})

# --- EXPORT TO EXCEL ---

@login_required
@user_passes_test(is_manager, login_url='pos_screen')
def export_excel_report(request):
    orders = Order.objects.all().order_by('-order_date')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Professional Styles
    title_font = Font(bold=True, size=16, color="2E5077")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5077", end_color="2E5077", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # Header Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "RESTAURANT SALES SUMMARY REPORT"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Table Headers
    headers = ['Order ID', 'Date & Time', 'Cashier', 'Subtotal', 'GST (5%)', 'Service (10%)', 'Discount', 'Total Amount']
    ws.append(headers)

    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data Rows
    for o in orders:
        ws.append([
            o.id, 
            o.order_date.strftime("%Y-%m-%d %H:%M"), 
            o.cashier.username, 
            float(o.subtotal), 
            float(o.gst_amount), 
            float(o.service_charge), 
            float(o.discount_amount), 
            float(o.total_amount)
        ])

    # Fixed: Adjust Column Widths (Skips Merged Title)
    for col in ws.columns:
        max_length = 0
        column_letter = col[1].column_letter # Use the header row to get the letter
        for cell in col:
            try:
                # Do not calculate width for merged cells
                if not isinstance(cell, openpyxl.cell.merged.MergedCell):
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Sales_Report.xlsx'
    wb.save(response)
    return response

# --- EXPORT TO PDF ---

@login_required
def export_pdf_bill(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    items = OrderItem.objects.filter(order=order)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Header / Branding
    p.setFillColor(colors.HexColor("#2E5077"))
    p.rect(0, height-100, width, 100, fill=1)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 24)
    p.drawCentredString(width/2, height-50, "POS PRO RESTAURANT")
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, height-75, "Official Tax Invoice | GST IN: 22AAAAA0000A1Z5")

    # Order Info
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(0.5*inch, height-130, f"INVOICE NO: #{order.id}")
    p.setFont("Helvetica", 10)
    p.drawString(0.5*inch, height-145, f"Date: {order.order_date.strftime('%d %b %Y, %I:%M %p')}")
    p.drawString(0.5*inch, height-160, f"Cashier: {order.cashier.username.upper()}")

    # Table Header Line
    p.setStrokeColor(colors.lightgrey)
    y = height - 190
    p.line(0.5*inch, y, 7.8*inch, y)
    p.setFont("Helvetica-Bold", 10)
    y -= 15
    p.drawString(0.6*inch, y, "ITEM DESCRIPTION")
    p.drawString(4.5*inch, y, "QTY")
    p.drawString(5.5*inch, y, "PRICE")
    p.drawString(6.8*inch, y, "TOTAL")
    y -= 10
    p.line(0.5*inch, y, 7.8*inch, y)

    # Table Items
    y -= 20
    p.setFont("Helvetica", 10)
    for oi in items:
        p.drawString(0.6*inch, y, f"{oi.item.name}")
        p.drawString(4.6*inch, y, f"{oi.quantity}")
        p.drawString(5.5*inch, y, f"Rs. {oi.price_at_order}")
        p.drawString(6.8*inch, y, f"Rs. {oi.quantity * oi.price_at_order}")
        y -= 20
        if y < 1.5*inch:
            p.showPage()
            y = height - 50

    # Summary Section
    y -= 10
    p.line(4.5*inch, y, 7.8*inch, y)
    y -= 20
    p.drawString(5.0*inch, y, "Subtotal:")
    p.drawRightString(7.7*inch, y, f"Rs. {order.subtotal:.2f}")
    y -= 15
    p.drawString(5.0*inch, y, "GST (5%):")
    p.drawRightString(7.7*inch, y, f"Rs. {order.gst_amount:.2f}")
    y -= 15
    p.drawString(5.0*inch, y, "Service (10%):")
    p.drawRightString(7.7*inch, y, f"Rs. {order.service_charge:.2f}")
    
    if order.discount_amount > 0:
        y -= 15
        p.setFillColor(colors.darkgreen)
        p.drawString(5.0*inch, y, "Discount Applied:")
        p.drawRightString(7.7*inch, y, f"- Rs. {order.discount_amount:.2f}")
        p.setFillColor(colors.black)

    y -= 25
    p.setFont("Helvetica-Bold", 14)
    p.drawString(5.0*inch, y, "GRAND TOTAL:")
    p.drawRightString(7.7*inch, y, f"Rs. {order.total_amount:.2f}")

    # Footer
    p.setFont("Helvetica-Oblique", 9)
    p.drawCentredString(width/2, 0.5*inch, "Thank you for dining with us! Please come again.")

    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'Invoice_{order.id}.pdf')